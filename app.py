from flask import Flask, render_template, request, send_file, redirect, url_for, session
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pdfplumber
import io
from datetime import timedelta
from psf_dashboard import create_dash_app



# Flask-configuratie
app = Flask(__name__)

dash_app = create_dash_app(app)
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"pdf"}
app.secret_key = "volvomanuel"

# Stel de sessie in om permanent te zijn
app.permanent_session_lifetime = timedelta(hours=1)  # Sessie blijft  min

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Zorg ervoor dat de uploads-map bestaat
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Globale buffer voor het gegenereerde Excel-bestand
output_buffer = None


def allowed_file(filename):
    """Controleer of het bestandstype toegestaan is."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# Middleware om te controleren of de intro bekeken is
@app.before_request
def require_intro():
    if "intro_viewed" not in session and request.endpoint not in ["intro", "static"]:
        return redirect(url_for("intro"))


# Route voor de intro-pagina
@app.route("/")
def intro():
    session["intro_viewed"] = True  # Geen session.permanent
    return render_template("intro.html")


# Route voor de homepagina
@app.route("/home")
def home():
    return render_template("home.html")
    
#Route voor PSF dashboard
from psf_dashboard import parse_excel, generate_plot
import plotly.io as pio

global_psf_data = {}

@app.route("/psf-dashboard", methods=["GET", "POST"])
def psf_dashboard():
    chart_html = None
    timers = []
    npts = []

    if request.method == 'POST':
        file = request.files.get('excel_file')
        if file:
            content = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64.b64encode(file.read()).decode()}"
            df = parse_excel(content)
            global_psf_data['df'] = df

        df = global_psf_data.get('df')
        if df is not None:
            timer = request.form.get('timer') or None
            npt = request.form.get('npt') or None
            nok_only = request.form.get('nok_only') == 'on'
            sigma_filter = request.form.get('sigma_filter') == 'on'
            min_welds = int(request.form.get('min_welds') or 0)
            max_welds = int(request.form.get('max_welds') or 9999)

            fig = generate_plot(df, timer, npt, min_welds, max_welds, nok_only, sigma_filter)
            chart_html = pio.to_html(fig, full_html=False)

            timers = df['TimerName'].dropna().unique()
            npts = df['NPTName'].dropna().unique()

    return render_template("psf.html", chart_html=chart_html, timers=timers, npts=npts)



@app.route("/bom-converter")
def index():
    """Startpagina."""
    return render_template("bom-converter.html")


@app.route("/autobom")
def autobom():
    """Startpagina."""
    return render_template("autobom.html")


@app.route("/upload", methods=["POST"])
def upload_files():
    """Verwerkt uploads en biedt een downloadlink aan."""
    global output_buffer  # Gebruik de globale buffer om het bestand op te slaan
    if "files[]" not in request.files:
        return "Geen bestanden geüpload", 400

    files = request.files.getlist("files[]")
    pdf_paths = []

    for file in files:
        if file and allowed_file(file.filename):
            pdf_paths.append(file)

    if pdf_paths:
        # Verwerk PDF's naar een Excel-bestand
        output_buffer = process_multiple_pdfs(pdf_paths)
        # Toon de bevestigingspagina
        return render_template("confirmation.html")

    return redirect(url_for("index"))


def process_single_pdf(pdf_file):
    """Verwerkt één PDF-bestand en zet data om naar een DataFrame."""
    with pdfplumber.open(pdf_file) as pdf:
        all_data = []
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                all_data.extend(table)

    # Zet data om naar DataFrame
    df = pd.DataFrame(all_data)
    if df.empty or len(df) <= 5:
        return pd.DataFrame()

    # Begin vanaf rij 5
    df = df.iloc[5:].reset_index(drop=True)

    # Rijen met de onderstaande woord niet overkopiëren

    df = df[
        ~df.apply(
            lambda row: row.astype(str).str.contains("COUNTER ELECTRODE").any(), axis=1
        )
    ]

    # Maak een nieuw DataFrame met de juiste structuur
    new_df = pd.DataFrame()
    new_df["H"] = df[1]  # B → H
    new_df["I"] = df[6]  # G → I
    new_df["P"] = df[6]  # G → P
    new_df["Z"] = df[6]  # G → Z
    new_df["M"] = df[8]  # I → M
    new_df["N"] = df[9]  # J → N
    new_df["Q"] = df[12]  # M → Q
    new_df["R"] = df[14]  # O → R
    new_df["Y"] = df[14]  # O → Y
    new_df["S"] = df[15]  # P → S

    # Speciale regels voor de eerste rij
    if not new_df.empty:
        g_value = new_df.at[0, "P"]
        r_value = new_df.at[0, "R"]
        combined_value = f"{g_value} {r_value}"
        new_df.at[0, "P"] = combined_value
        new_df.at[0, "Z"] = f"COMPLETE {combined_value}"
        new_df.at[0, "I"] = ""  # Kolom I blijft leeg
        new_df.at[0, "S"] = ""  # Kolom S blijft leeg

    return new_df


def process_multiple_pdfs(pdf_files):
    """Verwerkt meerdere PDF-bestanden en combineert ze in één Excel-bestand."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    # Kleurinstellingen
    yellow_fill = PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    )
    blue_fill = PatternFill(start_color="83CCEB", end_color="83CCEB", fill_type="solid")
    highlight_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )

    # Verwerk elk bestand
    for pdf_file in pdf_files:
        processed_df = process_single_pdf(pdf_file)
        if not processed_df.empty:
            # Stel de naam van de sheet in
            sheet_name = os.path.splitext(os.path.basename(pdf_file.filename))[0][:31]
            sheet_name = sheet_name.replace("volvo", "").strip()  # Verwijder volvo
            sheet_name = sheet_name[:31]  # 31 tekens max

            ws = workbook.create_sheet(title=sheet_name)

            # Kolom A startwaarden
            value_in_a = 10

            # Schrijf gegevens naar de sheet
            for r_idx, row in enumerate(processed_df.itertuples(index=False), start=1):
                # Kolom A: Schrijf waarde en pas kleur toe
                ws.cell(row=r_idx, column=1, value=value_in_a)
                if r_idx == 1:
                    ws.cell(row=r_idx, column=1).fill = yellow_fill  # Eerste rij geel
                else:
                    ws.cell(row=r_idx, column=1).fill = blue_fill  # Andere rijen blauw

                if r_idx > 1:
                    value_in_a += 10  # Verhoog met 10 na de eerste rij

                # Kolommen H tot Z
                ws.cell(row=r_idx, column=8, value=row[0])  # H
                ws.cell(row=r_idx, column=9, value=row[1])  # I
                ws.cell(row=r_idx, column=16, value=row[2])  # P
                ws.cell(row=r_idx, column=26, value=row[3])  # Z
                ws.cell(row=r_idx, column=13, value=row[4])  # M
                ws.cell(row=r_idx, column=14, value=row[5])  # N
                ws.cell(row=r_idx, column=17, value=row[6])  # Q
                ws.cell(row=r_idx, column=18, value=row[7])  # R
                ws.cell(row=r_idx, column=25, value=row[8])  # Y
                ws.cell(row=r_idx, column=19, value=row[9])  # S

                # Kleur kolommen Y en Z geel
                ws.cell(row=r_idx, column=25).fill = highlight_fill  # Y
                ws.cell(row=r_idx, column=26).fill = highlight_fill  # Z

    # Sla het bestand op in een BytesIO-buffer
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.route("/download")
def download_file():
    """Stelt het geconverteerde bestand beschikbaar voor download."""
    global output_buffer
    if output_buffer is None:
        return "Geen bestand beschikbaar om te downloaden", 400

    return send_file(
        output_buffer,
        as_attachment=True,
        download_name="converted-bom.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))  # Railway gebruikt de dynamische poort
    app.run(debug=False, host="0.0.0.0", port=port)




