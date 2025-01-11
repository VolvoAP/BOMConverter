from tkinter import Tk, Label, Button, filedialog, StringVar, Text, Scrollbar, END, Frame, Entry
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
import os
import warnings
import copy


warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Constants
MOVING_ARM_KEYWORD = "Moving Arm"
FIXED_ARM_KEYWORD = "Fixed Arm"
CONSOLE_KEYWORD = "Console"
# Tab color definitions
TAB_COLOR_COMPLETE = Color(rgb="FF00FF00")  # Green for complete ALLE 3 SUCCESVOL
TAB_COLOR_PARTIAL = Color(rgb="FF800080")   # PAARS GEDEELTELIJK (ARM)
TAB_COLOR_MISSING = Color(rgb="FFFF0000")   # Red (CONSOLE ISSUE)



LOG_FILE = "process_log.txt"

# Global variables
equipment_file = None
converted_files = []
main_file = None
log_text_widget = None  # Reference to the GUI text widget for logs

def update_log(message):
    """Update the log in the GUI text widget and save it to a file."""
    if log_text_widget:
        log_text_widget.insert(END, message + "\n")
        log_text_widget.see(END)  # Scroll to the end
    with open(LOG_FILE, "a") as log_file:
        log_file.write(message + "\n")

def download_log():
    """Download the current log file to the local machine."""
    download_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    if download_path:
        with open(LOG_FILE, "r") as log_file:
            log_content = log_file.read()
        with open(download_path, "w") as output_file:
            output_file.write(log_content)
        update_log(f"Log downloaded to {download_path}")        

def clear_log_file():
    """Clear the log file at the start of the application."""
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, "w") as log_file:
            log_file.write("=== Log Cleared on Startup ===\n")

def search_log():
    """Search for the text in the log and highlight matches."""
    search_term = search_entry.get()
    log_text_widget.tag_remove("highlight", "1.0", END)  # Remove existing highlights

    if not search_term:
        return

    start_pos = "1.0"
    matches = 0

    while True:
        start_pos = log_text_widget.search(search_term, start_pos, stopindex=END)
        if not start_pos:
            break

        end_pos = f"{start_pos}+{len(search_term)}c"
        log_text_widget.tag_add("highlight", start_pos, end_pos)
        matches += 1
        start_pos = end_pos

    if matches > 0:
        update_log(f"Found {matches} match(es) for '{search_term}'.")
        log_text_widget.tag_config("highlight", background="yellow", foreground="black")
    else:
        update_log(f"No matches found for '{search_term}'.")            

def log_summary(log, processed_count, skipped_count, duplicate_count):
    update_log(f"=== Processing Summary ===\n")
    update_log(f"Total Rows Processed: {processed_count + skipped_count}\n")
    update_log(f"Rows Skipped: {skipped_count} (Empty/N/A)\n")
    update_log(f"Rows Successfully Processed: {processed_count}\n")
    update_log(f"Duplicated Locations Skipped: {duplicate_count}\n")
    update_log("\n")

def log_location_details(log, location, moving_status, fixed_status, console_status):
    update_log(f"- Location {location}:\n")
    update_log(f"    * Moving Arm: {moving_status}\n")
    update_log(f"    * Fixed Arm: {fixed_status}\n")
    update_log(f"    * Console: {console_status}\n")

def select_equipment_file(label_var):
    global equipment_file
    equipment_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    label_var.set(f"Selected: {os.path.basename(equipment_file)}" if equipment_file else "No file selected")


def select_converted_files(label_var):
    global converted_files
    converted_files = filedialog.askopenfilenames(filetypes=[("Excel files", "*.xlsx")])
    if converted_files:
        filenames = [os.path.basename(file) for file in converted_files]
        label_var.set(f"Selected: {', '.join(filenames)}")
    else:
        label_var.set("No files selected")


def select_main_file(label_var):
    global main_file
    main_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    label_var.set(f"Selected: {os.path.basename(main_file)}" if main_file else "No file selected")




def find_section_row(target_sheet, section_keyword, log):
    """
    Find the start row of a section in the target sheet.
    """
    log.write(f"Searching for section start with keyword '{section_keyword}'.\n")
    for row in range(1, target_sheet.max_row + 1):
        cell_value = target_sheet.cell(row=row, column=7).value  # Assume keyword is in column G
        if isinstance(cell_value, str) and section_keyword.lower() in cell_value.lower():
            log.write(f"Found section '{section_keyword}' start at row {row}.\n")
            return row
    log.write(f"Section '{section_keyword}' not found.\n")
    return None


def remove_empty_rows(sheet, start_row, max_blank_rows=1):
    """
    Remove all rows in the section until the next G column value.
    Leaves only up to `max_blank_rows` between sections.
    """
    current_row = start_row + 1
    while current_row <= sheet.max_row:
        g_value = sheet.cell(row=current_row, column=7).value  # Check G column
        if g_value:  # Stop at the next G column value
            break
        row_values = [sheet.cell(row=current_row, column=col).value for col in range(1, sheet.max_column + 1)]
        if not any(row_values):  # Only remove empty rows
            sheet.delete_rows(current_row)
        else:
            current_row += 1

    # Ensure `max_blank_rows` after removing extra blank rows
    blank_row_count = 0
    while current_row <= sheet.max_row:
        row_values = [sheet.cell(row=current_row, column=col).value for col in range(1, sheet.max_column + 1)]
        if any(row_values):  # Stop at a non-blank row
            break
        blank_row_count += 1
        if blank_row_count > max_blank_rows:
            sheet.delete_rows(current_row)
        else:
            current_row += 1

def copy_data_exactly(source_sheet, target_sheet, target_start_row, log):
    """
    Copy all data from the source sheet to the target sheet starting at a specific row.
    """
    log.write(f"Starting copy from source sheet '{source_sheet.title}' to target sheet '{target_sheet.title}' at row {target_start_row}.\n")

    # Copy row by row
    for source_row in source_sheet.iter_rows(min_row=1, max_row=source_sheet.max_row):
        target_row_index = target_start_row + source_row[0].row - 1
        for source_cell in source_row:
            target_cell = target_sheet.cell(row=target_row_index, column=source_cell.column)

            # Copy value and styles
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)

    log.write(f"Finished copying data from {source_sheet.title} to {target_sheet.title}.\n")


    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        new_range = str(merged_cell_range)
        adjusted_range = new_range.replace(
            f"{source_sheet.title}!", 
            f"{target_sheet.title}!", 
        )
        target_sheet.merge_cells(adjusted_range)

    log.write(f"Finished exact copy from source sheet to target sheet starting at row {target_start_row}.\n")



def copy_data_to_section(source_sheet, target_sheet, section_start, log):
    """
    Copy data from the source sheet to the target sheet, ensuring space availability
    and respecting the next G-column value as a boundary.
    """
    log.write(f"Starting copy operation to section at row {section_start + 1}.\n")

    # Find the next G-column value or the end of the sheet
    current_row = section_start + 1
    while current_row <= target_sheet.max_row:
        g_value = target_sheet.cell(row=current_row, column=7).value  # Assuming column G is column 7
        if g_value:
            log.write(f"Boundary found at row {current_row}. G-column value: {g_value}.\n")
            break
        current_row += 1
    else:
        current_row = target_sheet.max_row + 1  # End of the sheet

    stop_row = current_row
    rows_available = stop_row - section_start - 1
    rows_needed = source_sheet.max_row  # Rows in the source sheet

    log.write(f"Rows available: {rows_available}. Rows needed: {rows_needed}.\n")

    # Adjust rows if needed
    if rows_needed > rows_available:
        rows_to_add = rows_needed - rows_available
        log.write(f"Adding {rows_to_add} rows to accommodate data.\n")
        target_sheet.insert_rows(stop_row, rows_to_add)
        stop_row += rows_to_add
    elif rows_available > rows_needed:
        rows_to_remove = rows_available - rows_needed
        log.write(f"Removing {rows_to_remove} extra rows to match data size.\n")
        target_sheet.delete_rows(stop_row - rows_to_remove, rows_to_remove)

    # Copy data from source to target
    log.write(f"Copying {rows_needed} rows from source to target starting at row {section_start + 1}.\n")
    for source_row_idx, source_row in enumerate(source_sheet.iter_rows(min_row=1, max_row=rows_needed), start=section_start + 1):
        for source_cell in source_row:
            target_cell = target_sheet.cell(row=source_row_idx, column=source_cell.column)

            # Copy value
            target_cell.value = source_cell.value

            # Copy styles
            if source_cell.has_style:
                target_cell.font = copy.copy(source_cell.font)
                target_cell.border = copy.copy(source_cell.border)
                target_cell.fill = copy.copy(source_cell.fill)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy.copy(source_cell.protection)
                target_cell.alignment = copy.copy(source_cell.alignment)
                log.write(f"Copied cell ({source_cell.row}, {source_cell.column}) -> ({target_cell.row}, {target_cell.column}). Value: {source_cell.value}\n")

    log.write(f"Finished copying {rows_needed} rows to section at row {section_start + 1}.\n")


def process_column(row, target_sheet, log):
    """
    Process a single row from the equipment file.
    Handles Moving Arm, Fixed Arm, and Console processing logic dynamically.
    """
    global converted_files

    # Extract MachineNumber and validate it starts with '5'
    machine_number = str(row['MachineNumber']).strip()
    if not machine_number.startswith('5'):
        log.write(f"[SKIP] Machine {machine_number} does not start with '5'. Skipping.\n")
        return False

    # Initialize status trackers for tab coloring
    success_count = 0
    total_columns = 3  # Completemov.Arm, Completefix.Arm, Console
    console_na_flag = False  # Track if the Console section is missing

    # Define processing logic for each column
    for column_name, section_keyword, is_console in [
        ('Completemov.Arm', 'Moving Arm', False),
        ('Completefix.Arm', 'Fixed Arm', False),
        ('Console', 'Console', True),
    ]:
        value = row[column_name]

        # Skip empty values or 'N/A'
        if is_na_value(value):
            log.write(f"[SKIP] {section_keyword}: Value is empty or marked as 'n/a' for Machine {machine_number}.\n")
            section_start = find_section_row(target_sheet, section_keyword, log)
            if section_start:
                clear_section_rows(target_sheet, section_start, target_sheet.max_row, log)
                log.write(f"[CLEAR] {section_keyword} section cleared for Machine {machine_number}.\n")

            # Specifically track if the Console section is missing
            if is_console:
                log.write(f"[CONSOLE] Console section is marked as 'N/A' for Machine {machine_number}. Setting console_na_flag to True.\n")
                console_na_flag = True
            continue

        # Process valid values
        value = str(value).strip()
        log.write(f"[PROCESS] Processing {section_keyword} for value {value} and Machine {machine_number}.\n")
        match_found = False
        for converted_file in converted_files:
            converted_wb = load_workbook(converted_file)
            if value in converted_wb.sheetnames:
                match_found = True
                source_sheet = converted_wb[value]

                # Find the start and stop rows in the target sheet
                section_start = find_section_row(target_sheet, section_keyword, log)
                if not section_start:
                    log.write(f"[FAIL] Section {section_keyword} start not found for Machine {machine_number}.\n")
                    break

                # Adjust rows and copy data
                log.write(f"Adjusting rows and copying data for {section_keyword}.\n")
                adjust_rows_and_copy(source_sheet, target_sheet, section_start, log)
                success_count += 1
                break

        if not match_found:
            log.write(f"[FAIL] Value {value} not found for {section_keyword} for Machine {machine_number}.\n")

    # Final tab color adjustment
    if console_na_flag:
        log.write(f"[TAB COLOR] Setting tab color to RED (missing) for Machine {machine_number} due to Console.\n")
        target_sheet.sheet_properties.tabColor = TAB_COLOR_MISSING  # Red if Console is missing
        return False
    
    elif success_count < total_columns:
        log.write(f"[TAB COLOR] Setting tab color to YELLOW (partial) for Machine {machine_number} due to Arm issues.\n")
        target_sheet.sheet_properties.tabColor = TAB_COLOR_PARTIAL  # Yellow if Arm issues
    else:
        log.write(f"[TAB COLOR] Setting tab color to GREEN (complete) for Machine {machine_number}.\n")
        target_sheet.sheet_properties.tabColor = TAB_COLOR_COMPLETE  # Green if all sections are processed

    log.write(f"[SUMMARY] Processing completed for Machine {machine_number}. {success_count}/{total_columns} successful.\n")
    return success_count == total_columns



def adjust_rows_and_copy(source_sheet, target_sheet, start_row, log):
    """
    Adjust rows in the target sheet dynamically and copy data from the source sheet.
    Ensure no data is overwritten and maintain the structure between G values.
    """
    log.write(f"Adjusting rows dynamically and copying data starting at row {start_row}.\n")

    # Zoek de huidige en volgende 'G' waarde
    current_g_row = start_row
    next_g_row = None

    # Zoek naar de volgende G waarde
    for row_idx in range(current_g_row + 1, target_sheet.max_row + 1):
        if target_sheet.cell(row=row_idx, column=7).value:  # Kolom G = 7
            next_g_row = row_idx
            break

    if not next_g_row:
        log.write(f"[WARNING] No next G value found after row {current_g_row}. Assuming end of sheet.\n")
        next_g_row = target_sheet.max_row + 1  # Geen volgende waarde, ga naar einde

    # Bereken beschikbare ruimte
    available_rows = next_g_row - current_g_row - 1
    needed_rows = source_sheet.max_row

    log.write(f"Available rows: {available_rows}, Needed rows: {needed_rows}.\n")

    # Verwijder bestaande data in het bereik
    log.write(f"Clearing existing data between rows {current_g_row + 1} and {next_g_row - 1}.\n")
    for row_idx in range(current_g_row + 1, next_g_row):
        for col_idx in range(1, target_sheet.max_column + 1):
            target_sheet.cell(row=row_idx, column=col_idx).value = None

    # Pas rijen aan indien nodig
    if needed_rows > available_rows:
        rows_to_add = needed_rows - available_rows
        log.write(f"Adding {rows_to_add} extra rows to fit data.\n")
        target_sheet.insert_rows(next_g_row, rows_to_add)
        next_g_row += rows_to_add

    elif needed_rows < available_rows:
        rows_to_remove = available_rows - needed_rows
        log.write(f"Removing {rows_to_remove} excess rows.\n")
        target_sheet.delete_rows(current_g_row + 1 + needed_rows, rows_to_remove)

    # Kopieer data naar de juiste positie
    log.write(f"Copying data from source to target between row {current_g_row + 1} and {next_g_row}.\n")
    copy_data_exactly(source_sheet, target_sheet, current_g_row + 1, log)

    log.write("Finished adjusting rows and copying data.\n")




def is_na_value(value):
    """
    Controleer of een waarde gelijk is aan 'n/a' of vergelijkbare waarden.
    """
    NA_VALUES = {"n/a", "n-a", "niet beschikbaar", "nvt", "n.v.t.","N/A"}  # Voeg meer synoniemen toe
    if value is None:
        return False
    if isinstance(value, str) and value.strip().lower() in NA_VALUES:
        return True
    return False




def process_column(row, column_name, section_keyword, target_sheet, log, is_console=False):
    """
    Process and copy data for a specific column (section).
    Handles both Arms and Console logic, with dynamic row adjustment.
    """
    global converted_files
    machine_number = str(row['MachineNumber']).strip()

    # Get the value from the column
    value = row[column_name]

    # Skip empty or 'N/A' values
    if pd.isna(value) or is_na_value(value):
        log.write(f"Skipping {section_keyword}: Value is empty or marked as 'n/a' for Machine {machine_number}.\n")
        return False

    # Convert to string and clean up
    if isinstance(value, float):
        value = str(int(value))
    else:
        value = str(value).strip()

    # Process valid values
    log.write(f"Processing {section_keyword}: Looking for value {value}.\n")
    for converted_file in converted_files:
        converted_wb = load_workbook(converted_file)
        if value in converted_wb.sheetnames:
            log.write(f"Match found for {section_keyword} in sheet {value}.\n")
            source_sheet = converted_wb[value]

            # Find the start of the section in the target sheet
            section_start = find_section_row(target_sheet, section_keyword, log)
            if not section_start:
                log.write(f"No section start found for {section_keyword}. Skipping.\n")
                return False

            log.write(f"Adjusting rows and copying data for {section_keyword}.\n")
            adjust_rows_and_copy(source_sheet, target_sheet, section_start, log)
            return True

    # Handle partial matches or no matches
    log.write(f"Value {value} not found for {section_keyword}.\n")
    return False




def clear_section_rows(sheet, start_row, end_row, log):
    """
    Clear all rows between start_row and end_row in the target sheet.
    """
    log.write(f"Clearing rows from {start_row} to {end_row - 1}.\n")
    for row_idx in range(start_row, end_row):
        for col_idx in range(1, sheet.max_column + 1):
            sheet.cell(row=row_idx, column=col_idx).value = None


def fit_and_copy_data(source_sheet, target_sheet, start_row, section_keyword, log):
    """
    Ensure data from source fits dynamically in the target sheet without overwriting the next section.
    """
    log.write(f"Preparing to copy data for section '{section_keyword}' from source to target.\n")

    # Calculate required rows
    source_rows = list(source_sheet.iter_rows(values_only=True))
    rows_needed = len(source_rows)

    # Find the next G value row
    next_g_row = None
    for row_idx in range(start_row, target_sheet.max_row + 1):
        if target_sheet.cell(row=row_idx, column=7).value:  # Kolom G = 7
            next_g_row = row_idx
            break

    if not next_g_row:
        next_g_row = target_sheet.max_row + 1

    available_rows = next_g_row - start_row
    log.write(f"Rows needed: {rows_needed}, available: {available_rows}.\n")

    # Clear existing data in the section
    log.write(f"Clearing data between rows {start_row} and {next_g_row - 1}.\n")
    clear_section_rows(target_sheet, start_row, next_g_row, log)

    # Adjust rows if needed
    if rows_needed > available_rows:
        log.write(f"Insufficient rows available. Adding {rows_needed - available_rows} rows.\n")
        target_sheet.insert_rows(next_g_row, rows_needed - available_rows)

    # Copy the data
    log.write(f"Copying {rows_needed} rows of data to target sheet.\n")
    copy_data_exactly(source_sheet, target_sheet, start_row, log)


def remove_section(sheet, start_row, log):
    """
    Remove a section starting at a specific row, including the G value row,
    until the next G column value is found.
    """
    log.write(f"Removing rows starting from row {start_row} in target sheet.\n")

    # Remove the G value row itself
    sheet.delete_rows(start_row)
    log.write(f"Deleted G value row at {start_row}.\n")

    # Continue removing rows until the next G column value
    current_row = start_row
    while current_row <= sheet.max_row:
        g_value = sheet.cell(row=current_row, column=7).value  # Check G column
        if g_value:  # Stop at the next G value
            break
        sheet.delete_rows(current_row)
        log.write(f"Deleted row {current_row}.\n")

    # Ensure one blank row remains
    sheet.insert_rows(current_row)
    log.write(f"Inserted one blank row at row {current_row}.\n")

''''
def remove_console_section(target_sheet, section_start, log):
    """
    Remove a console section in the target sheet and clean up empty rows.
    """
    current_row = section_start + 1
    while current_row <= target_sheet.max_row:
        g_value = target_sheet.cell(row=current_row, column=7).value
        if g_value:
            break
        current_row += 1

    stop_row = current_row  # First row with a G value
    log.write(f"Removing rows from {section_start + 1} to {stop_row} in target sheet.\n")
    target_sheet.delete_rows(section_start + 1, stop_row - section_start)

    # Clean up any remaining empty rows
    remove_empty_rows(target_sheet, section_start, max_blank_rows=1)
    target_sheet.insert_rows(section_start + 1, 1)  # Add one empty row
'''


'''
def process_console(row, column_name, target_sheet, log):
    """
    Process console logic: Skip, copy, or remove based on value in the Console column.
    """
    value = row[column_name]
    location = row["Location"]  # Kolom A in de Equipment-bestand
    if pd.isna(value):
        log.write(f"Skipping Console: Value is empty for location {location}.\n")
        return False

    # Controleer of de waarde N/A is
    if isinstance(value, str) and value.strip().lower() in ["n/a", "N/A"]:
        log.write(f"Console is marked as N/A for location {location}. Removing Console section.\n")

        # Zoek locatie in Main-bestand
        target_location = f"GA-{location}"
        if target_location not in target_sheet.sheetnames:
            log.write(f"Location {target_location} not found in main file.\n")
            return False

        target_ws = target_sheet[target_location]

        # Zoek naar Console in kolom G
        section_start = find_section_row(target_ws, "console", log)
        if section_start:
            remove_console_section(target_ws, section_start, log)
        else:
            log.write(f"Console section not found in {target_location}.\n")
        return True

    # Anders, waarde verwerken
    log.write(f"Processing Console: Value '{value}' for location {location}.\n")
    return process_column(row, column_name, "Console", target_sheet, log, is_console=True)
'''

def process_equipment():
    """Main function to process the equipment file."""
    global equipment_file, converted_files, main_file

    if not all([equipment_file, converted_files, main_file]):
        update_log("Please select all required files.")
        return

    print("Processing started...")
    update_log("Processing started...")
    with open(LOG_FILE, "w") as log:
        log.write("=== Processing Started ===\n")
        equipment_df = pd.read_excel(equipment_file)
        equipment_df = equipment_df[equipment_df['MachineNumber'].str.startswith('5', na=False)]
        main_wb = load_workbook(main_file)

        for _, row in equipment_df.iterrows():
            machine_number = row['MachineNumber']
            target_sheet_name = f"GA-{machine_number}"

            if target_sheet_name not in main_wb.sheetnames:
                log.write(f"Location {target_sheet_name}: Not found. Skipping.\n")
                continue

            target_sheet = main_wb[target_sheet_name]

            # Process each section and track results
            success_moving = process_column(row, 'Completemov.Arm', MOVING_ARM_KEYWORD, target_sheet, log)
            success_fixed = process_column(row, 'Completefix.Arm', FIXED_ARM_KEYWORD, target_sheet, log)
            success_console = process_column(row, 'CONSOLE', CONSOLE_KEYWORD, target_sheet, log, is_console=True)

            # Summarize the results for this location
            update_log(f"Location {target_sheet_name}:\n")
            update_log(f"- Moving Arm: {'Processed' if success_moving else 'Skipped'}\n")
            update_log(f"- Fixed Arm: {'Processed' if success_fixed else 'Skipped'}\n")
            update_log(f"- Console: {'Processed' if success_console else 'Skipped'}\n\n")

            # Tab color rules
            if not success_console:
                target_sheet.sheet_properties.tabColor = TAB_COLOR_MISSING
                continue

            elif success_moving and success_fixed and success_console:
                target_sheet.sheet_properties.tabColor = TAB_COLOR_COMPLETE
            else:
                target_sheet.sheet_properties.tabColor = TAB_COLOR_PARTIAL

        main_wb.save(main_file)
        update_log ("Processing complete. Check log file for details.")

    print("Processing complete. Check log file for details.")


def create_gui():
    global log_text_widget, search_entry

    root = Tk()
    root.title("BOM Processor")

    # Stel de venstergrootte in en plaats het venster in het midden
    window_width = 800
    window_height = 600
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = (screen_width // 2) - (window_width // 2)
    position_y = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")

    # Main frame to center everything
    main_frame = Frame(root)
    main_frame.pack(expand=True)

    # Frame voor bestandselectie
    file_frame = Frame(main_frame)
    file_frame.pack(pady=10)

    equipment_label = StringVar(value="No file selected")
    converted_label = StringVar(value="No files selected")
    main_label = StringVar(value="No file selected")

    # File selection labels and buttons
    Label(file_frame, text="Select Equipment File:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    Button(file_frame, text="Browse", command=lambda: select_equipment_file(equipment_label)).grid(row=0, column=1, padx=5)
    Label(file_frame, textvariable=equipment_label).grid(row=0, column=2, sticky="w", padx=5)

    Label(file_frame, text="Select Converted BOM Files:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    Button(file_frame, text="Browse", command=lambda: select_converted_files(converted_label)).grid(row=1, column=1, padx=5)
    Label(file_frame, textvariable=converted_label).grid(row=1, column=2, sticky="w", padx=5)

    Label(file_frame, text="Select Main BOM File:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    Button(file_frame, text="Browse", command=lambda: select_main_file(main_label)).grid(row=2, column=1, padx=5)
    Label(file_frame, textvariable=main_label).grid(row=2, column=2, sticky="w", padx=5)

    # Start processing button
    Button(file_frame, text="Start Processing", command=process_equipment).grid(row=3, column=1, pady=10)

    # Search bar for log
    search_frame = Frame(main_frame)
    search_frame.pack(pady=5)

    search_entry = Entry(search_frame, width=50)
    search_entry.pack(side="left", padx=5)

    Button(search_frame, text="Search", command=search_log).pack(side="left")

    # Download log button
    Button(search_frame, text="Download Log", command=download_log).pack(side="left", padx=5)

    # Text widget for log display
    log_frame = Frame(main_frame)
    log_frame.pack(fill="both", expand=True, padx=20, pady=10)

    log_text_widget = Text(log_frame, wrap="word")
    log_text_widget.pack(side="left", fill="both", expand=True)

    # Scrollbar for the log text widget
    scrollbar = Scrollbar(log_frame, command=log_text_widget.yview)
    scrollbar.pack(side="right", fill="y")
    log_text_widget.config(yscrollcommand=scrollbar.set)

    # Clear the log file at startup
    clear_log_file()

    root.mainloop()

if __name__ == "__main__":
    process_equipment()
